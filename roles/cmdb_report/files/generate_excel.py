#!/usr/bin/env python3
import json
import os
import sys
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Le module openpyxl n'est pas installé. Tentative d'installation...")
    try:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        print("openpyxl a été installé avec succès!")
    except Exception as e:
        print(f"Échec de l'installation de openpyxl: {e}")
        print("Assurez-vous que le package 'python3-venv' est installé sur votre système.")
        print("Conseil: Exécutez 'sudo apt install python3-venv python3-full' avant de relancer le playbook.")
        sys.exit(1)

def load_config():
    config_path = sys.argv[1] if len(sys.argv) > 1 else 'report_config.json'
    try:
        with open(config_path, 'r') as f:
            return json.load(f)
    except Exception as e:
        print(f"Erreur lors du chargement de la configuration: {e}")
        sys.exit(1)

def load_data(config):
    data = []
    data_dir = config.get('data_dir', 'json')
    
    # Vérifier si le répertoire existe
    if not os.path.exists(data_dir):
        print(f"ERREUR: Le répertoire {data_dir} n'existe pas!")
        sys.exit(1)
    
    # Vérifier si des fichiers JSON existent dans le répertoire
    json_files = [f for f in os.listdir(data_dir) if f.endswith('.json')]
    if not json_files:
        print(f"ERREUR: Aucun fichier JSON trouvé dans {data_dir}")
        sys.exit(1)
    
    # Charger les fichiers JSON
    for filename in json_files:
        try:
            with open(os.path.join(data_dir, filename), 'r') as f:
                data.append(json.load(f))
        except Exception as e:
            print(f"Erreur lors du chargement du fichier {filename}: {e}")
    
    return data

def format_header(ws, row=1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        cell.border = thin_border

def adjust_column_width(ws):
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 50)

def create_summary_sheet(wb, data, config):
    if not config.get('sheets', {}).get('summary', True):
        return
    
    ws = wb.create_sheet("Résumé")
    
    # Calculate counts by OS, environment and virtualization type
    os_count = {}
    env_count = {}
    virt_count = {}
    servers_with_updates = 0
    critical_servers = 0
    servers_with_expired_certs = 0
    
    for server in data:
        # Count by OS
        os_name = f"{server['software']['os']['distribution']} {server['software']['os']['distribution_version']}"
        os_count[os_name] = os_count.get(os_name, 0) + 1
        
        # Count by environment
        env = server['organizational']['environment']
        env_count[env] = env_count.get(env, 0) + 1
        
        # Count by virtualization
        virt = server['hardware']['system']['virtualization_type']
        virt_count[virt] = virt_count.get(virt, 0) + 1
        
        # Count servers with updates
        if isinstance(server['security'].get('updates_available'), int) and server['security'].get('updates_available', 0) > 0:
            servers_with_updates += 1
        
        # Count critical servers
        if server['organizational'].get('criticality') == 'Critique':
            critical_servers += 1
        
        # Check for expired certificates
        for cert in server['security'].get('certificates', {}).get('details', []):
            if 'stdout' in cert and 'notAfter' in cert['stdout']:
                expiration_line = [line for line in cert.get('stdout_lines', []) if 'notAfter' in line][0] if 'stdout_lines' in cert else ""
                if expiration_line and '2024' in expiration_line:  # Simplified check for expired certs in 2024
                    servers_with_expired_certs += 1
                    break
    
    # Add summary info
    ws['A1'] = "RÉSUMÉ DE L'INVENTAIRE CMDB"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    
    ws['A3'] = "Rapport généré le:"
    ws['B3'] = datetime.now().strftime(config.get('date_format', '%d/%m/%Y %H:%M:%S'))
    
    ws['A4'] = "Nombre total de serveurs:"
    ws['B4'] = len(data)
    
    ws['A5'] = "Serveurs nécessitant des mises à jour:"
    ws['B5'] = servers_with_updates
    
    ws['A6'] = "Serveurs critiques:"
    ws['B6'] = critical_servers
    
    ws['A7'] = "Serveurs avec certificats expirés:"
    ws['B7'] = servers_with_expired_certs
    
    # Distribution par OS
    row = 10
    ws['A9'] = "DISTRIBUTION PAR SYSTÈME D'EXPLOITATION"
    ws['A9'].font = Font(bold=True)
    ws.merge_cells('A9:C9')
    
    ws['A' + str(row)] = "Système d'exploitation"
    ws['B' + str(row)] = "Nombre"
    ws['C' + str(row)] = "Pourcentage"
    format_header(ws, row)
    
    row += 1
    for os, count in sorted(os_count.items(), key=lambda x: x[1], reverse=True):
        ws['A' + str(row)] = os
        ws['B' + str(row)] = count
        ws['C' + str(row)] = f"{count / len(data) * 100:.1f}%"
        row += 1
    
    # Distribution par environnement
    row += 2
    ws['A' + str(row)] = "DISTRIBUTION PAR ENVIRONNEMENT"
    ws['A' + str(row)].font = Font(bold=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws['A' + str(row)] = "Environnement"
    ws['B' + str(row)] = "Nombre"
    ws['C' + str(row)] = "Pourcentage"
    format_header(ws, row)
    
    row += 1
    for env, count in sorted(env_count.items(), key=lambda x: x[1], reverse=True):
        ws['A' + str(row)] = env
        ws['B' + str(row)] = count
        ws['C' + str(row)] = f"{count / len(data) * 100:.1f}%"
        row += 1
    
    # Distribution par type de virtualisation
    row += 2
    ws['A' + str(row)] = "DISTRIBUTION PAR TYPE DE VIRTUALISATION"
    ws['A' + str(row)].font = Font(bold=True)
    ws.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws['A' + str(row)] = "Type de virtualisation"
    ws['B' + str(row)] = "Nombre"
    ws['C' + str(row)] = "Pourcentage"
    format_header(ws, row)
    
    row += 1
    for virt, count in sorted(virt_count.items(), key=lambda x: x[1], reverse=True):
        ws['A' + str(row)] = virt
        ws['B' + str(row)] = count
        ws['C' + str(row)] = f"{count / len(data) * 100:.1f}%"
        row += 1
    
    adjust_column_width(ws)
    return ws

def create_servers_sheet(wb, data, config):
    if not config.get('sheets', {}).get('servers', True):
        return
    
    ws = wb.create_sheet("Serveurs")
    
    # Create header
    headers = [
        "Hostname", "FQDN", "Distribution", "Version", "Kernel", 
        "IP Principale", "Environnement", "Rôle", "Criticité",
        "Applications", "Date de collecte"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    format_header(ws)
    
    # Fill data
    for row, server in enumerate(data, 2):
        ws.cell(row=row, column=1).value = server.get('hostname', 'N/A')
        ws.cell(row=row, column=2).value = server.get('network', {}).get('fqdn', 'N/A')
        ws.cell(row=row, column=3).value = server.get('software', {}).get('os', {}).get('distribution', 'N/A')
        ws.cell(row=row, column=4).value = server.get('software', {}).get('os', {}).get('distribution_version', 'N/A')
        ws.cell(row=row, column=5).value = server.get('software', {}).get('kernel', {}).get('name', 'N/A')
        
        # Get default IP if available
        if server.get('network', {}).get('default_ipv4', {}).get('address'):
            ws.cell(row=row, column=6).value = server['network']['default_ipv4']['address']
        else:
            # Try to get the first interface IP
            interfaces = server.get('network', {}).get('interfaces', [])
            if interfaces and interfaces[0].get('ipv4', {}).get('address'):
                ws.cell(row=row, column=6).value = interfaces[0]['ipv4']['address']
            else:
                ws.cell(row=row, column=6).value = "N/A"
        
        ws.cell(row=row, column=7).value = server.get('organizational', {}).get('environment', 'N/A')
        ws.cell(row=row, column=8).value = server.get('organizational', {}).get('role', 'N/A')
        ws.cell(row=row, column=9).value = server.get('organizational', {}).get('criticality', 'N/A')
        
        # Format applications list
        apps = server.get('organizational', {}).get('applications', [])
        ws.cell(row=row, column=10).value = ", ".join(apps) if apps else "N/A"
        
        # Format collection date
        col_date = server.get('collection_date')
        if col_date:
            try:
                # Handle different date formats (with or without T/Z)
                if 'T' in col_date:
                    date_format = '%Y-%m-%dT%H:%M:%SZ' if col_date.endswith('Z') else '%Y-%m-%dT%H:%M:%S'
                    dt = datetime.strptime(col_date, date_format)
                else:
                    dt = datetime.strptime(col_date, '%Y-%m-%d %H:%M:%S')
                ws.cell(row=row, column=11).value = dt.strftime(config.get('date_format', '%d/%m/%Y %H:%M:%S'))
            except:
                ws.cell(row=row, column=11).value = col_date
        else:
            ws.cell(row=row, column=11).value = "N/A"
    
    adjust_column_width(ws)
    return ws

def create_hardware_sheet(wb, data, config):
    if not config.get('sheets', {}).get('hardware', True):
        return
    
    ws = wb.create_sheet("Matériel")
    
    # Create header
    headers = [
        "Hostname", "Modèle", "Fabricant", "Numéro de série",
        "Architecture", "Type de virtualisation", "Rôle de virtualisation",
        "Processeurs", "Cœurs", "Threads par cœur", "Mémoire (MB)", "Disques"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    format_header(ws)
    
    # Fill data
    for row, server in enumerate(data, 2):
        hw = server.get('hardware', {})
        
        ws.cell(row=row, column=1).value = server.get('hostname', 'N/A')
        ws.cell(row=row, column=2).value = hw.get('system', {}).get('model', 'N/A')
        ws.cell(row=row, column=3).value = hw.get('system', {}).get('manufacturer', 'N/A')
        ws.cell(row=row, column=4).value = hw.get('system', {}).get('serial', 'N/A')
        ws.cell(row=row, column=5).value = hw.get('system', {}).get('architecture', 'N/A')
        ws.cell(row=row, column=6).value = hw.get('system', {}).get('virtualization_type', 'N/A')
        ws.cell(row=row, column=7).value = hw.get('system', {}).get('virtualization_role', 'N/A')
        ws.cell(row=row, column=8).value = hw.get('processor', {}).get('count', 'N/A')
        ws.cell(row=row, column=9).value = hw.get('processor', {}).get('cores', 'N/A')
        ws.cell(row=row, column=10).value = hw.get('processor', {}).get('threads_per_core', 'N/A')
        ws.cell(row=row, column=11).value = hw.get('memory', {}).get('total_mb', 'N/A')
        
        # Format disks info
        disks = hw.get('disks', [])
        disk_info = []
        for disk in disks:
            if 'name' in disk and 'size' in disk:
                disk_info.append(f"{disk['name']}:{disk['size']}")
        ws.cell(row=row, column=12).value = ", ".join(disk_info) if disk_info else "N/A"
    
    adjust_column_width(ws)
    return ws

def create_software_sheet(wb, data, config):
    if not config.get('sheets', {}).get('software', True):
        return
    
    ws = wb.create_sheet("Logiciels")
    
    # Create header
    headers = [
        "Hostname", "Distribution", "Version", "Release", 
        "Kernel", "Python Version", "Packages installés", 
        "Services en cours d'exécution", "Bases de données"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    format_header(ws)
    
    # Fill data
    for row, server in enumerate(data, 2):
        sw = server.get('software', {})
        
        ws.cell(row=row, column=1).value = server.get('hostname', 'N/A')
        ws.cell(row=row, column=2).value = sw.get('os', {}).get('distribution', 'N/A')
        ws.cell(row=row, column=3).value = sw.get('os', {}).get('distribution_version', 'N/A')
        ws.cell(row=row, column=4).value = sw.get('os', {}).get('distribution_release', 'N/A')
        ws.cell(row=row, column=5).value = sw.get('kernel', {}).get('name', 'N/A')
        ws.cell(row=row, column=6).value = sw.get('python', {}).get('version', 'N/A')
        
        # Package count
        pkgs = sw.get('packages', {})
        if isinstance(pkgs, dict) and 'count' in pkgs:
            ws.cell(row=row, column=7).value = pkgs['count']
        else:
            ws.cell(row=row, column=7).value = "N/A"
        
        # Running services
        services = sw.get('services', {}).get('running', [])
        ws.cell(row=row, column=8).value = ", ".join(services) if services else "N/A"
        
        # Databases
        dbs = sw.get('databases', [])
        db_info = []
        for db in dbs:
            if 'type' in db and 'version' in db:
                db_info.append(f"{db['type']} {db['version'].split()[0]}")
        ws.cell(row=row, column=9).value = ", ".join(db_info) if db_info else "N/A"
    
    adjust_column_width(ws)
    return ws

def create_network_sheet(wb, data, config):
    if not config.get('sheets', {}).get('network', True):
        return
    
    ws = wb.create_sheet("Réseau")
    
    # Create header
    headers = [
        "Hostname", "FQDN", "Domaine", "IPv4 Principale", "Masque",
        "Passerelle", "Interfaces", "DNS Serveurs", "DNS Search"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    format_header(ws)
    
    # Fill data
    for row, server in enumerate(data, 2):
        net = server.get('network', {})
        
        ws.cell(row=row, column=1).value = server.get('hostname', 'N/A')
        ws.cell(row=row, column=2).value = net.get('fqdn', 'N/A')
        ws.cell(row=row, column=3).value = net.get('domain', 'N/A')
        
        # Default IPv4
        default_ipv4 = net.get('default_ipv4', {})
        ws.cell(row=row, column=4).value = default_ipv4.get('address', 'N/A')
        ws.cell(row=row, column=5).value = default_ipv4.get('netmask', 'N/A')
        ws.cell(row=row, column=6).value = default_ipv4.get('gateway', 'N/A')
        
        # Interfaces
        interfaces = net.get('interfaces', [])
        if_info = []
        for iface in interfaces:
            if 'name' in iface and 'mac' in iface:
                if_addr = iface.get('ipv4', {}).get('address', '')
                if_info.append(f"{iface['name']}:{if_addr}")
        ws.cell(row=row, column=7).value = ", ".join(if_info) if if_info else "N/A"
        
        # DNS Configuration
        dns = net.get('dns', {})
        ws.cell(row=row, column=8).value = ", ".join(dns.get('nameservers', [])) if dns.get('nameservers') else "N/A"
        ws.cell(row=row, column=9).value = ", ".join(dns.get('search', [])) if dns.get('search') else "N/A"
    
    adjust_column_width(ws)
    return ws

def create_security_sheet(wb, data, config):
    if not config.get('sheets', {}).get('security', True):
        return
    
    ws = wb.create_sheet("Sécurité")
    
    # Create header
    headers = [
        "Hostname", "SELinux", "AppArmor", "Mises à jour disponibles",
        "Utilisateurs shell", "Utilisateurs root", "Utilisateurs sudo",
        "Certificats", "Solutions de sauvegarde"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    format_header(ws)
    
    # Fill data
    for row, server in enumerate(data, 2):
        sec = server.get('security', {})
        
        ws.cell(row=row, column=1).value = server.get('hostname', 'N/A')
        ws.cell(row=row, column=2).value = sec.get('selinux', {}).get('status', 'N/A')
        ws.cell(row=row, column=3).value = sec.get('apparmor', {}).get('status', 'N/A')
        
        # Updates available
        updates = sec.get('updates_available', 'N/A')
        ws.cell(row=row, column=4).value = updates
        if isinstance(updates, int) and updates > 0:
            ws.cell(row=row, column=4).fill = PatternFill(start_color='FFAAAA', end_color='FFAAAA', fill_type='solid')
        
        # Users
        users = sec.get('users', {})
        ws.cell(row=row, column=5).value = users.get('with_shell', 'N/A')
        ws.cell(row=row, column=6).value = users.get('with_uid0', 'N/A')
        ws.cell(row=row, column=7).value = users.get('with_sudo', 'N/A')
        
        # Certificates
        certs = sec.get('certificates', {})
        ws.cell(row=row, column=8).value = certs.get('count', 0)
        
        # Backup solutions
        backup = sec.get('backup', {})
        backup_solutions = backup.get('solutions', [])
        ws.cell(row=row, column=9).value = ", ".join(backup_solutions) if backup_solutions else "Aucune détectée"
    
    adjust_column_width(ws)
    return ws

def create_organizational_sheet(wb, data, config):
    if not config.get('sheets', {}).get('organizational', True):
        return
    
    ws = wb.create_sheet("Organisation")
    
    # Create header
    headers = [
        "Hostname", "Rôle", "Environnement", "Criticité",
        "Applications", "Responsable métier", "Responsable technique",
        "SLA", "Date de mise en service", "Date de fin de vie",
        "Datacenter", "Rack", "Position"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    format_header(ws)
    
    # Fill data
    for row, server in enumerate(data, 2):
        org = server.get('organizational', {})
        
        ws.cell(row=row, column=1).value = server.get('hostname', 'N/A')
        ws.cell(row=row, column=2).value = org.get('role', 'N/A')
        ws.cell(row=row, column=3).value = org.get('environment', 'N/A')
        ws.cell(row=row, column=4).value = org.get('criticality', 'N/A')
        
        # Format applications list
        apps = org.get('applications', [])
        ws.cell(row=row, column=5).value = ", ".join(apps) if apps else "N/A"
        
        ws.cell(row=row, column=6).value = org.get('business_owner', 'N/A')
        ws.cell(row=row, column=7).value = org.get('technical_owner', 'N/A')
        ws.cell(row=row, column=8).value = org.get('sla', 'N/A')
        ws.cell(row=row, column=9).value = org.get('commissioned_date', 'N/A')
        ws.cell(row=row, column=10).value = org.get('end_of_life_date', 'N/A')
        ws.cell(row=row, column=11).value = org.get('datacenter', 'N/A')
        ws.cell(row=row, column=12).value = org.get('rack', 'N/A')
        ws.cell(row=row, column=13).value = org.get('position', 'N/A')
    
    adjust_column_width(ws)
    return ws

def create_certificates_sheet(wb, data, config):
    if not config.get('sheets', {}).get('certificates', True):
        return
    
    ws = wb.create_sheet("Certificats")
    
    # Create header
    headers = [
        "Hostname", "Sujet", "Émetteur", "Valide depuis", "Valide jusqu'à", "Chemin"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    format_header(ws)
    
    # Fill data
    row = 2
    for server in data:
        hostname = server.get('hostname', 'N/A')
        
        cert_details = server.get('security', {}).get('certificates', {}).get('details', [])
        
        if not cert_details:
            # Add a row for servers with no certificates
            ws.cell(row=row, column=1).value = hostname
            ws.cell(row=row, column=2).value = "Aucun certificat trouvé"
            row += 1
            continue
        
        for cert in cert_details:
            if 'stdout' not in cert or not cert['stdout']:
                continue
                
            ws.cell(row=row, column=1).value = hostname
            
            # Parse certificate details
            subject = ""
            issuer = ""
            valid_from = ""
            valid_to = ""
            path = cert.get('item', {}).get('path', 'N/A') if 'item' in cert else 'N/A'
            
            if 'stdout_lines' in cert:
                for line in cert['stdout_lines']:
                    if 'subject=' in line:
                        subject = line.replace('subject=', '')
                    elif 'issuer=' in line:
                        issuer = line.replace('issuer=', '')
                    elif 'notBefore=' in line:
                        valid_from = line.replace('notBefore=', '')
                    elif 'notAfter=' in line:
                        valid_to = line.replace('notAfter=', '')
            
            ws.cell(row=row, column=2).value = subject
            ws.cell(row=row, column=3).value = issuer
            ws.cell(row=row, column=4).value = valid_from
            ws.cell(row=row, column=5).value = valid_to
            ws.cell(row=row, column=6).value = path
            
            # Highlight expired or soon-to-expire certificates
            if valid_to and '2024' in valid_to:
                ws.cell(row=row, column=5).fill = PatternFill(start_color='FFAAAA', end_color='FFAAAA', fill_type='solid')
            
            row += 1
    
    adjust_column_width(ws)
    return ws

def create_updates_sheet(wb, data, config):
    if not config.get('sheets', {}).get('updates', True):
        return
    
    ws = wb.create_sheet("Mises à jour")
    
    # Create header
    headers = [
        "Hostname", "Distribution", "Version", "Mises à jour disponibles", "SELinux", "AppArmor"
    ]
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header
    format_header(ws)
    
    # Fill data
    row = 2
    for server in data:
        ws.cell(row=row, column=1).value = server.get('hostname', 'N/A')
        ws.cell(row=row, column=2).value = server.get('software', {}).get('os', {}).get('distribution', 'N/A')
        ws.cell(row=row, column=3).value = server.get('software', {}).get('os', {}).get('distribution_version', 'N/A')
        
        # Updates available
        updates = server.get('security', {}).get('updates_available', 'N/A')
        ws.cell(row=row, column=4).value = updates
        if isinstance(updates, int) and updates > 0:
            ws.cell(row=row, column=4).fill = PatternFill(start_color='FFAAAA', end_color='FFAAAA', fill_type='solid')
        
        # Security features
        ws.cell(row=row, column=5).value = server.get('security', {}).get('selinux', {}).get('status', 'N/A')
        ws.cell(row=row, column=6).value = server.get('security', {}).get('apparmor', {}).get('status', 'N/A')
        
        row += 1
    
    adjust_column_width(ws)
    return ws

def main():
    try:
        # Load configuration
        config = load_config()
        
        # Load data
        data = load_data(config)
        
        if not data:
            print("No data found! Please check the data directory.")
            sys.exit(1)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Create sheets based on configuration
        create_summary_sheet(wb, data, config)
        create_servers_sheet(wb, data, config)
        create_hardware_sheet(wb, data, config)
        create_software_sheet(wb, data, config)
        create_network_sheet(wb, data, config)
        create_security_sheet(wb, data, config)
        create_organizational_sheet(wb, data, config)
        create_certificates_sheet(wb, data, config)
        create_updates_sheet(wb, data, config)
        
        # Save the workbook
        output_file = config.get('output_file', 'cmdb_inventory_report.xlsx')
        wb.save(output_file)
        print(f"Report generated successfully: {output_file}")
        
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()